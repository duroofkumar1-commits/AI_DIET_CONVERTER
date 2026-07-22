"""
exporter.py
-----------
Generates downloadable reports (PDF via ReportLab, Excel via OpenPyXL,
CSV) summarising a meal's nutrition and its vegetarian recommendations.
"""

from __future__ import annotations

import csv
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                 TableStyle)

from utils.constants import APP_NAME, HEADLINE_NUTRIENTS
from utils.helpers import format_nutrient, pretty_nutrient_name

REPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")


def _ensure_reports_dir() -> None:
    os.makedirs(REPORTS_DIR, exist_ok=True)


def export_pdf(meal_foods: list[tuple[str, float]], target_nutrients: dict[str, float],
                recommendations: list[dict], filename: str | None = None) -> str:
    """Build a PDF report with the entered meal and top vegetarian recommendations."""
    _ensure_reports_dir()
    filename = filename or f"diet_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    path = os.path.join(REPORTS_DIR, filename)

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(path, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    story = [
        Paragraph(APP_NAME, styles["Title"]),
        Paragraph(f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')}", styles["Normal"]),
        Spacer(1, 14),
        Paragraph("Non-Vegetarian Meal Entered", styles["Heading2"]),
    ]

    meal_table_data = [["Food", "Quantity (g)"]] + [[name, f"{grams:.1f}"] for name, grams in meal_foods]
    story.append(_styled_table(meal_table_data))
    story.append(Spacer(1, 14))

    story.append(Paragraph("Total Nutrition Profile", styles["Heading2"]))
    nutrient_table_data = [["Nutrient", "Amount"]] + [
        [pretty_nutrient_name(k), format_nutrient(k, v)] for k, v in target_nutrients.items()
    ]
    story.append(_styled_table(nutrient_table_data))
    story.append(Spacer(1, 14))

    story.append(Paragraph("AI Vegetarian Recommendations", styles["Heading2"]))
    for i, rec in enumerate(recommendations, 1):
        story.append(Paragraph(
            f"Recommendation {i} &mdash; Overall Match: {rec['overall_match']}%",
            styles["Heading3"]))
        rec_table_data = [["Food", "Quantity (g)"]] + [
            [f["name"], f"{f['grams']:.1f}"] for f in rec["foods"]
        ]
        story.append(_styled_table(rec_table_data))
        headline_line = " | ".join(
            f"{pretty_nutrient_name(k)}: {v:.0f}%" for k, v in rec["headline_matches"].items()
        )
        story.append(Paragraph(headline_line, styles["Normal"]))
        story.append(Spacer(1, 10))

    doc.build(story)
    return path


def _styled_table(data: list[list[str]]) -> Table:
    table = Table(data, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2fa572")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))
    return table


def export_excel(meal_foods: list[tuple[str, float]], target_nutrients: dict[str, float],
                  recommendations: list[dict], filename: str | None = None) -> str:
    """Build an Excel workbook with meal, nutrition totals, and recommendations."""
    _ensure_reports_dir()
    filename = filename or f"diet_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(REPORTS_DIR, filename)

    wb = Workbook()
    header_fill = PatternFill(start_color="2FA572", end_color="2FA572", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws1 = wb.active
    ws1.title = "Meal & Nutrients"
    ws1.append(["Food", "Quantity (g)"])
    for cell in ws1[1]:
        cell.fill, cell.font = header_fill, header_font
    for name, grams in meal_foods:
        ws1.append([name, round(grams, 1)])
    ws1.append([])
    ws1.append(["Nutrient", "Amount"])
    for cell in ws1[ws1.max_row]:
        cell.fill, cell.font = header_fill, header_font
    for k, v in target_nutrients.items():
        ws1.append([pretty_nutrient_name(k), round(v, 2)])
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 22

    ws2 = wb.create_sheet("Recommendations")
    ws2.append(["Recommendation #", "Food", "Quantity (g)", "Overall Match %"])
    for cell in ws2[1]:
        cell.fill, cell.font = header_fill, header_font
    for i, rec in enumerate(recommendations, 1):
        for f in rec["foods"]:
            ws2.append([i, f["name"], f["grams"], rec["overall_match"]])
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 22

    wb.save(path)
    return path


def export_csv(meal_foods: list[tuple[str, float]], target_nutrients: dict[str, float],
               recommendations: list[dict], filename: str | None = None) -> str:
    """Plain CSV export (meal + nutrients + recommendations in one file)."""
    _ensure_reports_dir()
    filename = filename or f"diet_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    path = os.path.join(REPORTS_DIR, filename)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Section", "Food", "Quantity (g)"])
        for name, grams in meal_foods:
            writer.writerow(["Meal", name, round(grams, 1)])
        writer.writerow([])
        writer.writerow(["Nutrient", "Amount"])
        for k, v in target_nutrients.items():
            writer.writerow([pretty_nutrient_name(k), round(v, 2)])
        writer.writerow([])
        writer.writerow(["Recommendation #", "Food", "Quantity (g)", "Overall Match %"])
        for i, rec in enumerate(recommendations, 1):
            for f in rec["foods"]:
                writer.writerow([i, f["name"], f["grams"], rec["overall_match"]])
    return path
